# -*- coding: utf-8 -*-
"""
Created on Wed Nov 19 16:12:20 2025

@author: AMAURY
"""

from openpyxl import load_workbook, Workbook
from pathlib import Path
from datetime import date
import shutil

# === Rutas: ajusta si tus archivos están en otra carpeta ===
file_deps = Path(r"Reglas_R1_R22_dependencias_corregidas.xlsx")
file_resumen = Path(r"Resumen_impl_vs_excel.xlsx")

today_tag = "2025-11-18"  # usa fecha fija para reproducibilidad
sheet_estado = f"Estado_{today_tag}"
sheet_corr = "Correcciones_dependencias"

# -------- Datos de las dos pestañas --------
estado_rows = [
    ["Regla","Detector/Evento","Estado","Comentarios","Dependencias mínimas","Formato de salida"],
    ["R4","absorption","Hecho","Ventana deslizante dur_s; vol_btc alto; usa best bid/ask; drift reducido",
     "trades + best of book (depth)","texto: '#R4 BUY/SELL absorption @ px | vol=nnn BTC'"],
    ["R5","slicing_aggr=iceberg","Hecho","Slices agresivos del MISMO tamaño (require_equal=True); mismo price/side; gap_ms<=80",
     "trades","texto: '#R5 {BUY/SELL} slicing_iceberg @ px | qty=Σ BTC'"],
    ["R6","slicing_aggr=hitting","Hecho","Slices agresivos NO idénticos (require_equal=False); mismo price/side; gap_ms<=80",
     "trades","texto: '#R6 {BUY/SELL} slicing_hitting @ px | qty=Σ BTC'"],
    ["R7","slicing_pass","Hecho","INSERT consecutivos al mismo price/side; k_min configurable",
     "depth (insert)","texto: '#R7 {BUY/SELL} slicing_pass @ px | qty=Σ BTC'"],
    ["R8","break_wall","Hecho","Reacciona a slicing_aggr; gating con depleción/refill y |basis_vel_bps_s|; forbids refill≥60%/3s",
     "slicing_aggr + depth metrics + mark/index","texto: '#R8 {BUY/SELL} break_wall @ px | k=n'"],
    ["R9","dominance","Hecho","Dominancia por niveles (levels=1000 por defecto); spread≤max_spread_usd",
     "depth (head N)","texto: '#R9 {BUY/SELL} dominance xx.x% @ px'"],
    ["R10","dominance","Hecho","Mismo detector; el R-code lo decide eval_rules según lado/perfil",
     "depth (head N)","texto: '#R10 {BUY/SELL} dominance xx.x% @ px'"],
    ["R11","aggressor_imbalance","Pendiente","Ratio market buy/sell corto plazo con drift/vol gating",
     "trades (+ opcional best drift)","texto por definir"],
    ["R12","spoofing/passive-pull","Pendiente","Adds/pulls top-N con impacto",
     "depth","texto por definir"],
    ["R13","gamma_exposure_shift","Pendiente","Requiere opciones Deribit; GEX",
     "opciones (Deribit), mark/index","texto por definir"],
    ["R14","funding_anomaly","Pendiente","Outliers en funding y vel.",
     "mark_funding","texto por definir"],
    ["R15","open_interest_spike","Pendiente","Jump en OI con trades/price",
     "OI + trades","texto por definir"],
    ["R16","top_trader_ratios","Pendiente","Feed TTR de Binance",
     "TTR feed Binance","texto por definir"],
    ["R17","queue_position_risk","Pendiente","Riesgo por bursts de adds/cancels",
     "depth granular","texto por definir"],
    ["R18","cvd_divergence","Pendiente","CVD vs price",
     "trades, price","texto por definir"],
    ["R19","iceberg_reveal","Pendiente","Revelado tras pulls",
     "trades + depth pull/add","texto por definir"],
    ["R20","liquidation_sweep","Pendiente","Liq bursts",
     "liquidations + depth","texto por definir"],
    ["R21","basis_accel_break","Pendiente","Aceleración de basis",
     "mark/index (basis)","texto por definir"],
    ["R22","range_breakout_book","Pendiente","Ruptura de rango + orderbook",
     "trades + depth metrics","texto por definir"],
]

deps_rows = [
    ["Regla","Fuentes","Métricas/umbrales","Notas"],
    ["R4","trades + depth(best)","∑qty lado / dur_s ; best_bid/ask; drift opcional","max_price_drift_ticks se puede reactivar"],
    ["R5","trades","bucket same side+price; k_min; require_equal=True; equal_tol; qty_min","Iceberg (igualdad estricta)"],
    ["R6","trades","bucket same side+price; k_min; require_equal=False; qty_min","Hitting (sin igualdad)"],
    ["R7","depth","INSERT secuenciales al mismo price/side; k_min; qty_min","Colocación pasiva en lotes"],
    ["R8","slicing_aggr + depth metrics + mark/index","n_min; dep_pct; forbid_refill_under_pct (3s); |basis_vel_bps_s|≥thr","Usa snapshot de MetricsEngine"],
    ["R9/R10","depth (head N)","levels; dom_pct; max_spread_usd; hold_ms; retrigger_s","Dominancia por niveles no-cero"],
]

def backup(xlsx: Path) -> None:
    if xlsx.exists():
        shutil.copy2(xlsx, xlsx.with_suffix(".bak.xlsx"))

def write_new_sheet(xlsx_path: Path, sheet_name: str, rows: list[list]):
    backup(xlsx_path)
    if not xlsx_path.exists():
        wb = Workbook()
        # borra hoja por defecto si está vacía
        if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1:
            wb.remove(wb["Sheet"])
    else:
        wb = load_workbook(xlsx_path)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    ws = wb.create_sheet(title=sheet_name)
    for r in rows:
        ws.append(r)
    wb.save(xlsx_path)

# Ejecuta
write_new_sheet(file_resumen, sheet_estado, estado_rows)
write_new_sheet(file_deps, sheet_corr, deps_rows)

print(f"OK -> {file_resumen} [{sheet_estado}]")
print(f"OK -> {file_deps} [{sheet_corr}]")
